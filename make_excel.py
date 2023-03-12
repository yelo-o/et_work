import pandas as pd
import re

#회사 목록 불러오기
# from get_company_list import com_list  # 진짜 코드(테스트할 때는 닫아놓기)
com_list = ['KT&G', 'SK에너지']  # 가짜 변수(코드)


from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.drawing.image import Image

img = Image('stamp.png')
# A열의 값을 가져오기
for com in com_list:

    # 엑셀 파일
    wb = load_workbook(filename='22년 4분기 12월 매입매출정산서(비삼성)_샘플.xlsx')
    df = pd.read_excel('C:/flyordig/et_work/22년 4분기 12월 매입매출정산서(비삼성)_샘플.xlsx', sheet_name=2, engine='openpyxl')
    # print(wb.sheetnames)  # 파일의 시트를 불러옴

    '''
    # <1번 워크시트>
    print("1번시트")
    ws = wb['업무연락전']
    
    
    # 삭제할 회사 명단 저장
    list_to_delete = []  # 삭제해야할 행의 인덱스를 저장할 리스트 [0, 1, 2, ...]
    print(com)
    for row in ws.iter_rows(min_row=2, min_col=1):
        if row[0].value != com:
            if row[0].value != '#공통':
                # print(row[0].value,type(row[0].value))
                list_to_delete.append(row[0].row)
        else:
            pass
    
    list_to_delete.reverse()  # 하나씩 append했기 때문에 순서가 뒤바뀜


    print(f"1번 시트에서 업체가 {com}일 때 지워야할 줄 번호 목록은, {list_to_delete}는")
    print('행 갯수는' ,len(list_to_delete))
    
    # 삭제할 명단에 대해서 행 삭제 진행
    for row_index in list_to_delete:
        ws.delete_rows(row_index)  # 한 줄씩 삭제
    
    ## 엑셀 함수 재설정
    
    
    # D20
    formula = ws['D20'].value  # D20 셀에서 수식 가져오기
    formula_loc = formula[9:13]

    formula_changed = formula.replace(formula_loc,'$A20')
    print(f"기존 : {formula}, 변경 : {formula_changed} ")
    
    ws['D20'] = formula_changed
    
    # D21
    formula = ws['D21'].value  # D20 셀에서 수식 가져오기
    formula_loc = formula[9:13]

    formula_changed = formula.replace(formula_loc,'$A21')
    print(f"기존 : {formula}, 변경 : {formula_changed} ")
    
    ws['D21'] = formula_changed
    
    
    
    # D22
    formula = ws['D22'].value  # D20 셀에서 수식 가져오기
    formula_loc = formula[9:13]

    formula_changed = formula.replace(formula_loc,'$A22')
    print(f"기존 : {formula}, 변경 : {formula_changed} ")
    
    ws['D22'] = formula_changed
    
    cell = ws['D22']
    value = cell.value
    cell.value = value
    
    ws['D22'] = value
    
    
    # E20
    formula = ws['E20'].value  # D20 셀에서 수식 가져오기
    formula_loc = formula[9:13]

    formula_changed = formula.replace(formula_loc,'$A20')
    print(f"기존 : {formula}, 변경 : {formula_changed} ")
    
    ws['E20'] = formula_changed
    
    # E21
    formula = ws['E21'].value  # D20 셀에서 수식 가져오기
    formula_loc = formula[9:13]

    formula_changed = formula.replace(formula_loc,'$A21')
    print(f"기존 : {formula}, 변경 : {formula_changed} ")
    
    ws['E21'] = formula_changed
    
    
    # E22
    formula = ws['E22'].value  # D20 셀에서 수식 가져오기
    formula_loc = formula[9:13]

    formula_changed = formula.replace(formula_loc,'$A22')
    print(f"기존 : {formula}, 변경 : {formula_changed} ")
    
    ws['E22'] = formula_changed
    
    
    # F20
    ws['F20'] = '=D20+E20'
    # F21
    ws['F21'] = '=D21+E21'
    # F22
    ws['F22'] = '=D22+E22'
    # D23
    ws['D23'] = '=SUBTOTAL(9,D20:D22)'
    # E23
    ws['E23'] = '=SUBTOTAL(9,E20:E22)'
    # F23
    ws['F23'] = '=SUBTOTAL(9,F20:F22)'
    
    # F24
    formula = ws['F24'].value  # D20 셀에서 수식 가져오기
    formula_loc = formula[9:13]

    formula_changed = formula.replace(formula_loc,'$A24')
    print(f"기존 : {formula}, 변경 : {formula_changed} ")
    
    ws['F24'] = formula_changed
    
    # F25
    ws['F25'] = '=SUBTOTAL(9,F24:F24)'
    
    # F26
    ws['F26'] = '=F25+F23'
    
    # C17
    ws['C17'] = '=NUMBERSTRING(F26,1)&"원"&TEXT(F26,"(￦#,##0)")&"VAT별도"'
    
    # 병합
    
    ws.merge_cells("B19:C19")
    ws.merge_cells("B20:C20")
    ws.merge_cells("B21:C21")
    ws.merge_cells("B22:C22")
    ws.merge_cells("B23:C23")
    ws.merge_cells("B24:C24")
    ws.merge_cells("B25:C25")
    ws.merge_cells("B26:C26")
    ws.unmerge_cells("B38:C38") # 병합 해제 후
    ws.merge_cells("B38:G38") # 범위 재설정하여 재병합
    print("병합 완료")
    
    # 행 높이 조절
    ws.row_dimensions[11].height = 5.25
    ws.row_dimensions[16].height = 11.25
    ws.row_dimensions[19].height = 23
    ws.row_dimensions[20].height = 23
    ws.row_dimensions[21].height = 23
    ws.row_dimensions[22].height = 23
    ws.row_dimensions[23].height = 23
    ws.row_dimensions[24].height = 23
    ws.row_dimensions[25].height = 23
    
    ws.row_dimensions[26].height = 30
    
    ws.row_dimensions[27].height = 6.75
    ws.row_dimensions[28].height = 6
    ws.row_dimensions[28].height = 9.75
    
    ws.row_dimensions[30].height = 16.5
    ws.row_dimensions[32].height = 16.5
    ws.row_dimensions[33].height = 16.5
    ws.row_dimensions[34].height = 16.5
    ws.row_dimensions[35].height = 16.5
    ws.row_dimensions[36].height = 16.5
    ws.row_dimensions[37].height = 16.5
    ws.row_dimensions[38].height = 16.5
    ws.row_dimensions[39].height = 16.5
    ws.row_dimensions[40].height = 1
    
    print("행 높이 조절 완료")
    
    # 값 복사 붙여 넣기
    
    
    # A열 삭제
    # ws.delete_cols(1)
    
    # A열 숨기기
    ws.column_dimensions.group("A", hidden=True)
    
    # 인감 이미지 삽입
    ws.add_image(img, "F37")
    '''
    
    
    # <2번 워크시트>
    print("2번시트")
    ws = wb['거래명세표']  # 2번 시트 열기
    
    # 병합 임시 해제
    ws.unmerge_cells("I3:I25") # 공급자(I3:I25)
    for i in range(5,25,1):
        ws.unmerge_cells(f"L{i}:N{i}") # 회사주소
    for j in range(3,26,1):
        ws.unmerge_cells(f"J{j}:K{j}") # 업태
    
    
    
    # 행삭제 
    list_to_delete = []  # 삭제해야할 행의 인덱스를 저장할 리스트 [0, 1, 2, ...]
    for row in ws.iter_rows(min_row=2, min_col=1):
        if row[0].value != com:
            if row[0].value != '#공통':
                if row[0].value != '고객사':
                    list_to_delete.append(row[0].row)
        else:
            pass
        
    list_to_delete.reverse()  # 하나씩 append했기 때문에 순서가 뒤바뀜
    
    
    # 행 삭제 실행
    for row_index in list_to_delete:
        ws.delete_rows(row_index)  # 한 줄씩 삭제
        
    # 재병합
    ws.merge_cells("I3:I6")  # 공급자
    ws.merge_cells("L5:N5")  # 사업장 주소
    ws.merge_cells("D7:H7")  # 아래와 같이 계산합니다.
    ws.merge_cells("J3:K3")  # 등록번호
    ws.merge_cells("J4:K4")  # 상호
    ws.merge_cells("J5:K5")  # 사업장주소
    ws.merge_cells("J6:K6")  # 업태
    
    ## 엑셀 함수 재설정
    
    # C열
    last_row = ws.max_row
    while ws.cell(last_row, 1).value is None:
        last_row -= 1
    print("last_row : ", last_row)
        
    for row in range(9, last_row +1):
        
        formula = ws.cell(row, 3).value
        print("formula : ",formula)
        if formula is not None:
            pattern = r'(\$[A-Z]+\d+)'  # $A233    <- 이런식으로 받아오도록 정규식 표현
            print("pattern : ",pattern)
            match = re.search(pattern, formula)
            formula_loc = match.group(1)
            print("formula_loc : ", formula_loc)
            left_cell = ws.cell(row, 2).coordinate
            print("left_cell : ",left_cell)
            formula_changed = formula.replace(formula_loc, left_cell)
            print("formula_changed : ",formula_changed)
            ws[f'C{row}'] = formula_changed
    print("C열 마지막 행까지 함수 삽입 완료")    
    
    # D열
    for row in range(9, last_row +1):
        
        formula = ws.cell(row, 4).value
        print("formula : ",formula)
        if formula is not None:
            pattern = r'(\$[A-Z]+\d+)'  # $A233    <- 이런식으로 받아오도록 정규식 표현
            print("pattern : ",pattern)
            match = re.search(pattern, formula)
            formula_loc = match.group(1)
            print("formula_loc : ", formula_loc)
            left_cell = ws.cell(row, 2).coordinate
            print("left_cell : ",left_cell)
            formula_changed = formula.replace(formula_loc, left_cell)
            print("formula_changed : ",formula_changed)
            ws[f'D{row}'] = formula_changed
    print("D열 마지막 행까지 함수 삽입 완료")  
    
    # E열
    for row in range(9, last_row +1):
        
        formula = ws.cell(row, 5).value
        print("formula : ",formula)
        if formula is not None:
            pattern = r'(\$[A-Z]+\d+)'  # $A233    <- 이런식으로 받아오도록 정규식 표현
            print("pattern : ",pattern)
            match = re.search(pattern, formula)
            formula_loc = match.group(1)
            print("formula_loc : ", formula_loc)
            left_cell = ws.cell(row, 2).coordinate
            print("left_cell : ",left_cell)
            formula_changed = formula.replace(formula_loc, left_cell)
            print("formula_changed : ",formula_changed)
            ws[f'E{row}'] = formula_changed
    print("E열 마지막 행까지 함수 삽입 완료")  
    
    # F열
    for row in range(9, last_row +1):
        
        formula = ws.cell(row, 6).value
        print("formula : ",formula)
        if formula is not None:
            pattern = r'(\$[A-Z]+\d+)'  # $A233    <- 이런식으로 받아오도록 정규식 표현
            print("pattern : ",pattern)
            match = re.search(pattern, formula)
            formula_loc = match.group(1)
            print("formula_loc : ", formula_loc)
            left_cell = ws.cell(row, 2).coordinate
            print("left_cell : ",left_cell)
            formula_changed = formula.replace(formula_loc, left_cell)
            print("formula_changed : ",formula_changed)
            ws[f'F{row}'] = formula_changed
    print("F열 마지막 행까지 함수 삽입 완료")  
    
    # G열
    for row in range(9, last_row):
        
        formula = ws.cell(row, 7).value
        print("formula : ",formula)
        if formula is not None:
            pattern = r'(\$[A-Z]+\d+)'  # $A233    <- 이런식으로 받아오도록 정규식 표현
            print("pattern : ",pattern)
            match = re.search(pattern, formula)
            formula_loc = match.group(1)
            print("formula_loc : ", formula_loc)
            left_cell = ws.cell(row, 2).coordinate
            print("left_cell : ",left_cell)
            formula_changed = formula.replace(formula_loc, left_cell)
            print("formula_changed : ",formula_changed)
            ws[f'G{row}'] = formula_changed
    
    
    formula = ws.cell(last_row, 7).value
    print("formula : ",formula)
    pattern = r'([A-Z]+\d+:[A-Z]+\d+)'
    match = re.search(pattern, formula)
    formula_loc = match.group(1)
    upper_cell1 = ws.cell(9, 7).coordinate
    print("upper_cell1 : ", upper_cell1)
    upper_cell2 = ws.cell(last_row-1, 7).coordinate
    print("upper_cell2 : ", upper_cell2)
    upper_cell = f"{upper_cell1}:{upper_cell2}"
    print("upper_cell : ", upper_cell)
    formula_changed = formula.replace(formula_loc, upper_cell)
    print("formula_changed : ",formula_changed)
    ws[f'G{last_row}'] = formula_changed
    
    print("G열 Sub-total까지 함수 삽입 완료")
    # H열
    for row in range(9, last_row):
        
        formula = ws.cell(row, 8).value
        print("formula : ",formula)
        if formula is not None:
            pattern = r'(\$[A-Z]+\d+)'  # $A233    <- 이런식으로 받아오도록 정규식 표현
            print("pattern : ",pattern)
            match = re.search(pattern, formula)
            formula_loc = match.group(1)
            print("formula_loc : ", formula_loc)
            left_cell = ws.cell(row, 2).coordinate
            print("left_cell : ",left_cell)
            formula_changed = formula.replace(formula_loc, left_cell)
            print("formula_changed : ",formula_changed)
            ws[f'H{row}'] = formula_changed
    
    formula = ws.cell(last_row, 8).value
    print("formula : ",formula)
    pattern = r'([A-Z]+\d+:[A-Z]+\d+)'
    match = re.search(pattern, formula)
    formula_loc = match.group(1)
    upper_cell1 = ws.cell(9, 8).coordinate
    print("upper_cell1 : ", upper_cell1)
    upper_cell2 = ws.cell(last_row-1, 8).coordinate
    print("upper_cell2 : ", upper_cell2)
    upper_cell = f"{upper_cell1}:{upper_cell2}"
    print("upper_cell : ", upper_cell)
    formula_changed = formula.replace(formula_loc, upper_cell)
    print("formula_changed : ",formula_changed)
    ws[f'H{last_row}'] = formula_changed
    print("H열 Sub-total까지 함수 삽입 완료")
    
    # I열
    for row in range(9, last_row):
        
        formula = ws.cell(row, 9).value
        print("formula : ",formula)
        if formula is not None:
            pattern = r'(\$[A-Z]+\d+)'  # $A233    <- 이런식으로 받아오도록 정규식 표현
            print("pattern : ",pattern)
            match = re.search(pattern, formula)
            formula_loc = match.group(1)
            print("formula_loc : ", formula_loc)
            left_cell = ws.cell(row, 2).coordinate
            print("left_cell : ",left_cell)
            formula_changed = formula.replace(formula_loc, left_cell)
            print("formula_changed : ",formula_changed)
            ws[f'I{row}'] = formula_changed
            
            
    formula = ws.cell(last_row, 9).value
    print("formula : ",formula)
    pattern = r'([A-Z]+\d+:[A-Z]+\d+)'
    match = re.search(pattern, formula)
    formula_loc = match.group(1)
    upper_cell1 = ws.cell(9, 9).coordinate
    print("upper_cell1 : ", upper_cell1)
    upper_cell2 = ws.cell(last_row-1, 9).coordinate
    print("upper_cell2 : ", upper_cell2)
    upper_cell = f"{upper_cell1}:{upper_cell2}"
    print("upper_cell : ", upper_cell)
    formula_changed = formula.replace(formula_loc, upper_cell)
    print("formula_changed : ",formula_changed)
    ws[f'I{last_row}'] = formula_changed
    print("I열 Sub-total까지 함수 삽입 완료")
    
    # J열
    for row in range(9, last_row):
        
        formula = ws.cell(row, 10).value
        print("formula : ",formula)
        if formula is not None:
            pattern = r'(\$[A-Z]+\d+)'  # $A233    <- 이런식으로 받아오도록 정규식 표현
            print("pattern : ",pattern)
            match = re.search(pattern, formula)
            formula_loc = match.group(1)
            print("formula_loc : ", formula_loc)
            left_cell = ws.cell(row, 2).coordinate
            print("left_cell : ",left_cell)
            formula_changed = formula.replace(formula_loc, left_cell)
            print("formula_changed : ",formula_changed)
            ws[f'J{row}'] = formula_changed
            
    formula = ws.cell(last_row, 10).value
    print("formula : ",formula)
    pattern = r'([A-Z]+\d+:[A-Z]+\d+)'
    match = re.search(pattern, formula)
    formula_loc = match.group(1)
    upper_cell1 = ws.cell(9, 10).coordinate
    print("upper_cell1 : ", upper_cell1)
    upper_cell2 = ws.cell(last_row-1, 10).coordinate
    print("upper_cell2 : ", upper_cell2)
    upper_cell = f"{upper_cell1}:{upper_cell2}"
    print("upper_cell : ", upper_cell)
    formula_changed = formula.replace(formula_loc, upper_cell)
    print("formula_changed : ",formula_changed)
    ws[f'J{last_row}'] = formula_changed
    print("J열 Sub-total까지 함수 삽입 완료")
    
    # K열
    for row in range(9, last_row):
        
        formula = ws.cell(row, 11).value
        print("formula : ",formula)
        if formula is not None:
            pattern = r'(\$[A-Z]+\d+)'  # $A233    <- 이런식으로 받아오도록 정규식 표현
            print("pattern : ",pattern)
            match = re.search(pattern, formula)
            formula_loc = match.group(1)
            print("formula_loc : ", formula_loc)
            left_cell = ws.cell(row, 2).coordinate
            print("left_cell : ",left_cell)
            formula_changed = formula.replace(formula_loc, left_cell)
            print("formula_changed : ",formula_changed)
            ws[f'K{row}'] = formula_changed
            
    formula = ws.cell(last_row, 11).value
    print("formula : ",formula)
    pattern = r'([A-Z]+\d+:[A-Z]+\d+)'
    match = re.search(pattern, formula)
    formula_loc = match.group(1)
    upper_cell1 = ws.cell(9, 11).coordinate
    print("upper_cell1 : ", upper_cell1)
    upper_cell2 = ws.cell(last_row-1, 11).coordinate
    print("upper_cell2 : ", upper_cell2)
    upper_cell = f"{upper_cell1}:{upper_cell2}"
    print("upper_cell : ", upper_cell)
    formula_changed = formula.replace(formula_loc, upper_cell)
    print("formula_changed : ",formula_changed)
    ws[f'K{last_row}'] = formula_changed
    print("K열 Sub-total까지 함수 삽입 완료")       
            
    # M열

    for row in range(9, last_row):
        
        formula = ws.cell(row, 13).value  # M열은 13번째 열
        print("formula : ",formula)
        if formula is not None:
            pattern = r'[A-Z]\d+'  # $A233    <- 이런식으로 받아오도록 정규식 표현
            print("pattern : ",pattern)
            matches = re.findall(pattern, formula)
            print("matches : ", matches)
            left_cells = [ws.cell(row, 7).coordinate, ws.cell(row, 10).coordinate,
                          ws.cell(row, 11).coordinate, ws.cell(row, 12).coordinate,
                          ws.cell(row, 15).coordinate]
            print("left_cells : ",left_cells)
            for i in range(len(matches)):
                formula = formula.replace(matches[i], left_cells[i])
            print("참조값 변경 : ",formula)
            ws[f'M{row}'] = formula
            
    formula = ws.cell(last_row, 13).value
    print("formula : ",formula)
    pattern = r'([A-Z]+\d+:[A-Z]+\d+)'
    match = re.search(pattern, formula)
    formula_loc = match.group(1)
    upper_cell1 = ws.cell(9, 13).coordinate
    print("upper_cell1 : ", upper_cell1)
    upper_cell2 = ws.cell(last_row-1, 13).coordinate
    print("upper_cell2 : ", upper_cell2)
    upper_cell = f"{upper_cell1}:{upper_cell2}"
    print("upper_cell : ", upper_cell)
    formula_changed = formula.replace(formula_loc, upper_cell)
    print("formula_changed : ",formula_changed)
    ws[f'M{last_row}'] = formula_changed
    print("M열 Sub-total까지 함수 삽입 완료")      
    
    
            
    # 새로운 3번 시트 제작을 위해 기존 3번 시트 삭제
    # wb.remove(wb['구매내역'])
    
    # 새로운 4번 시트 제작을 위해 기존 4번 시트 삭제
    # wb.remove(wb['운송비'])
    
    
    '''
    
    # 변경된 내용 저장
    wb.save(f'{com}.xlsx')  # 파일 이름으로 저장
    print(f"{com} 파일 1차 저장")  
    
    # <3번 워크시트>
    print("3번시트")
    df = df[df.회사 == com]
    
    # 데이터프레임을 엑셀 파일 열어서 저장
    with pd.ExcelWriter(f'{com}.xlsx', mode='a', engine='openpyxl') as writer:
        writer.book = load_workbook(f'{com}.xlsx')
        df.to_excel(writer, sheet_name='구매내역', index=False)
    
    # 3번 시트 이쁘게 수정
    
    
    
    # <4번 워크시트>
    print("4번시트")
    
    # 3번 시트에서 초기화됐기 때문에 한번 더 불러줘야 함(df)
    df = pd.read_excel('C:/flyordig/et_work/22년 4분기 12월 매입매출정산서(비삼성)_샘플.xlsx', sheet_name=3, engine='openpyxl')
    df = df[df.회사 == com]
    
    # 데이터프레임을 엑셀 파일 열어서 저장
    with pd.ExcelWriter(f'{com}.xlsx', mode='a', engine='openpyxl') as writer:
        writer.book = load_workbook(f'{com}.xlsx')
        df.to_excel(writer, sheet_name='운송비', index=False)
    
    # 4번 시트 이쁘게 수정
    
    
    '''
    
    wb.save(f'{com}.xlsx')  # 파일 이름으로 저장
    print(f"{com} 파일 저장 및 완료") 
    
    
    
    
    

# 삭제할 행의 인덱스를 저장할 리스트
# 행 반복
# for row in ws.iter_rows(min_row=2, min_col=1, values_only=True):
#     # A열이 '현대중공업'이거나 '#공통'이 아닌 경우
#     if row[0] != '현대중공업' and row[0] != '#공통':
#         # 삭제할 행의 인덱스를 리스트에 추가
#         delete_rows.append(row[0].row)

# # 삭제할 행을 역순으로 정렬하여, 뒤에서부터 삭제하도록 함
# delete_rows.reverse()

# # 행 ws
# for row in delete_rows:
#     ws.delete_rows(row)

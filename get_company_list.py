from openpyxl import load_workbook

# 엑셀 파일
wb = load_workbook(filename='22년 4분기 12월 매입매출정산서(비삼성)_샘플.xlsx')
# print(wb.sheetnames)  # 파일의 시트를 불러옴

# 워크시트 선택
ws = wb['업무연락전']

# A열 변수 설정
list0 = []  # A열 데이터 읽어오기(중복포함)
list1 = []  # 중복 데이터 제거
com_list = []  # list1에서 None과 '#공통' 제거


# 변수 정리
for cell in ws['A']:
    list0.append(cell.value)
# print('list0은',list0, len(list0))

# list0에 있던 값들을 list1에 1개씩만 옮김(중복된 값은 append되지 않음)
for value in list0:
    if value not in list1:
        list1.append(value)
print('list1은',list1, len(list1))

# list1에서 None과 '#공통' 제거하여 com_list 만들기
com_list = list1
del com_list[0:2]
print('최종', com_list, len(com_list))

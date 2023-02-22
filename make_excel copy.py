import pandas as pd
from get_company_list import list0, list1, com_list
from openpyxl import load_workbook


# A열의 값을 가져오기
for com in com_list:

    # 엑셀 파일
    wb = load_workbook(filename='22년 4분기 12월 매입매출정산서(비삼성)_샘플.xlsx')
    # print(wb.sheetnames)  # 파일의 시트를 불러옴

    # 1번 워크시트
    print("1번시트")
    ws = wb['업무연락전']
    
    
    
    # 삭제할 회사 명단 저장
    list_to_delete = []  # 삭제해야할 행의 인덱스를 저장할 리스트 [0, 1, 2, ...]
    print(com)
    for row in ws.iter_rows(min_row=2, min_col=1):
        if row[0].value != com:
            if row[0].value != '#공통':
                # print(row[0].value,type(row[0].value))
                list_to_delete.append(row[0].row)
        else:
            pass
    
    list_to_delete.reverse()  # 하나씩 append했기 때문에 순서가 뒤바뀜


    print(f"1번 시트에서 업체가 {com}일 때 지워야할 줄 번호 목록은, {list_to_delete}는")
    print('행 갯수는' ,len(list_to_delete))
    
    # 삭제할 명단에 대해서 행 삭제 진행
    for row_index in list_to_delete:
        ws.delete_rows(row_index)  # 한 줄씩 삭제
        
        
        
    # 2번 워크시트
    print("2번시트")
    ws = wb['거래명세표']
    list_to_delete = []  # 삭제해야할 행의 인덱스를 저장할 리스트 [0, 1, 2, ...]
    for row in ws.iter_rows(min_row=2, min_col=1):
        if row[0].value != com:
            if row[0].value != '#공통':
                if row[0].value != '고객사':
                    list_to_delete.append(row[0].row)
        else:
            pass
        
    list_to_delete.reverse()  # 하나씩 append했기 때문에 순서가 뒤바뀜
    
    print(f"2번 시트에서 업체가 {com}일 때 지워야할 줄 번호 목록은, {list_to_delete}는")
    print('행 갯수는' ,len(list_to_delete))
    
    # 삭제할 명단에 대해서 행 삭제 진행
    for row_index in list_to_delete:
        ws.delete_rows(row_index)  # 한 줄씩 삭제
        
    
    
    
    
    
    # 3번 워크시트
    print("3번시트")
    ws = wb['구매내역']
    list_to_delete = []  # 삭제해야할 행의 인덱스를 저장할 리스트 [0, 1, 2, ...]
    for row in ws.iter_rows(min_row=2, min_col=1):
        if row[4].value != com:
            if row[4].value != '#공통':
                if row[4].value != '고객사':
                    list_to_delete.append(row[4].row)
        else:
            pass
        
    list_to_delete.reverse()  # 하나씩 append했기 때문에 순서가 뒤바뀜
    
    # print(f"3번 시트에서 업체가 {com}일 때 지워야할 줄 번호 목록은, {list_to_delete}는")
    print('행 갯수는' ,len(list_to_delete))
    
    # 삭제할 명단에 대해서 행 삭제 진행
    for row_index in list_to_delete:
        ws.delete_rows(row_index)  # 한 줄씩 삭제
    
       
    
    
    
    
    
    
    # 4번 워크시트
    
    
    
    # 변경된 내용 저장
    wb.save(f'{com}.xlsx')  # 파일 이름으로 저장
    print(f"{com} 파일 저장 및 완료")
    

# 삭제할 행의 인덱스를 저장할 리스트
# 행 반복
# for row in ws.iter_rows(min_row=2, min_col=1, values_only=True):
#     # A열이 '현대중공업'이거나 '#공통'이 아닌 경우
#     if row[0] != '현대중공업' and row[0] != '#공통':
#         # 삭제할 행의 인덱스를 리스트에 추가
#         delete_rows.append(row[0].row)

# # 삭제할 행을 역순으로 정렬하여, 뒤에서부터 삭제하도록 함
# delete_rows.reverse()

# # 행 ws
# for row in delete_rows:
#     ws.delete_rows(row)
